#!/usr/bin/env python3
"""Verify all downloaded XLS/XLSX files.

Detects:
- Redirect JSON stubs (should have been resolved by fix_xls_redirects.py)
- HTML error pages saved as Excel
- Corrupt or unreadable Excel files
- Empty workbooks
- Valid Excel files (counts sheets and rows)

Saves SHA256 checksums for all files.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from pathlib import Path
from typing import Any

import sys
sys.path.insert(0, str(Path(__file__).parent))

from junar_common import (
    ROOT,
    detect_junar_file_redirect,
    load_config,
    setup_logging,
    write_json,
)

AUDIT_DIR = ROOT / "audit"
MANIFEST_PATH = ROOT / "data/processed/download_manifest.json"

FILE_STATUSES = (
    "xlsx_valid",
    "xls_valid",
    "redirect_stub_unresolved",
    "html_error_file",
    "corrupt_excel",
    "empty_excel",
    "unsupported_legacy_xls",
)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def sniff_file_type(content_start: bytes) -> str:
    """Detect actual file type from magic bytes."""
    if content_start[:4] == b"PK\x03\x04":
        return "zip_or_xlsx"
    if content_start[:8] in (b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1",):
        return "ole_xls"
    if content_start[:5].lower() in (b"<html", b"<!doc"):
        return "html"
    if content_start[:1] in (b"{", b"["):
        return "json_like"
    return "unknown"


def inspect_excel(path: Path, logger) -> dict[str, Any]:
    ext = path.suffix.lower().lstrip(".")
    content_start = path.read_bytes()[:32]
    file_type = sniff_file_type(content_start)
    sha = sha256_file(path)

    # Check for redirect stub
    full_content = path.read_bytes()
    redirect = detect_junar_file_redirect(full_content)
    if redirect:
        return {
            "actual_file_type": "redirect_json",
            "sha256": sha,
            "sheet_count": None,
            "rows_total": None,
            "max_rows_single_sheet": None,
            "status": "redirect_stub_unresolved",
            "notes": f"fUri={redirect.get('fUri', '')[:80]}",
        }

    if file_type == "html":
        return {
            "actual_file_type": "html",
            "sha256": sha,
            "sheet_count": None,
            "rows_total": None,
            "max_rows_single_sheet": None,
            "status": "html_error_file",
            "notes": "",
        }

    if file_type == "json_like" and ext in ("xls", "xlsx"):
        return {
            "actual_file_type": "json",
            "sha256": sha,
            "sheet_count": None,
            "rows_total": None,
            "max_rows_single_sheet": None,
            "status": "html_error_file",
            "notes": "JSON content in Excel file",
        }

    # Try openpyxl for xlsx / zip
    if ext == "xlsx" or file_type == "zip_or_xlsx":
        try:
            import openpyxl
            # read_only=True does not populate max_row until fully iterated; use regular load.
            wb = openpyxl.load_workbook(path, data_only=True)
            sheets = wb.sheetnames
            sheet_rows = []
            for ws in wb.worksheets:
                # max_row includes header; subtract 1 for data rows.
                sheet_rows.append(max(0, (ws.max_row or 1) - 1))
            wb.close()
            rows_total = sum(sheet_rows)
            max_rows = max(sheet_rows) if sheet_rows else 0
            if rows_total == 0:
                status = "empty_excel"
            else:
                status = "xlsx_valid"
            return {
                "actual_file_type": "xlsx",
                "sha256": sha,
                "sheet_count": len(sheets),
                "rows_total": rows_total,
                "max_rows_single_sheet": max_rows,
                "status": status,
                "notes": f"sheets={sheets[:3]}",
            }
        except Exception as e:
            logger.warning("openpyxl failed on %s: %s", path.name, e)
            return {
                "actual_file_type": "xlsx_corrupt",
                "sha256": sha,
                "sheet_count": None,
                "rows_total": None,
                "max_rows_single_sheet": None,
                "status": "corrupt_excel",
                "notes": str(e)[:200],
            }

    # Try xlrd for old .xls
    if ext == "xls" or file_type == "ole_xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(path)
            sheet_rows = [max(0, ws.nrows - 1) for ws in wb.sheets()]
            rows_total = sum(sheet_rows)
            max_rows = max(sheet_rows) if sheet_rows else 0
            if rows_total == 0:
                status = "empty_excel"
            else:
                status = "xls_valid"
            return {
                "actual_file_type": "xls",
                "sha256": sha,
                "sheet_count": len(wb.sheets()),
                "rows_total": rows_total,
                "max_rows_single_sheet": max_rows,
                "status": status,
                "notes": "",
            }
        except xlrd.biffh.XLRDError as e:
            if "xlrd" in str(e).lower() or "xlsx" in str(e).lower():
                return {
                    "actual_file_type": "unsupported",
                    "sha256": sha,
                    "sheet_count": None,
                    "rows_total": None,
                    "max_rows_single_sheet": None,
                    "status": "unsupported_legacy_xls",
                    "notes": str(e)[:200],
                }
            return {
                "actual_file_type": "xls_corrupt",
                "sha256": sha,
                "sheet_count": None,
                "rows_total": None,
                "max_rows_single_sheet": None,
                "status": "corrupt_excel",
                "notes": str(e)[:200],
            }
        except Exception as e:
            return {
                "actual_file_type": "xls_error",
                "sha256": sha,
                "sheet_count": None,
                "rows_total": None,
                "max_rows_single_sheet": None,
                "status": "corrupt_excel",
                "notes": str(e)[:200],
            }

    return {
        "actual_file_type": file_type,
        "sha256": sha,
        "sheet_count": None,
        "rows_total": None,
        "max_rows_single_sheet": None,
        "status": "corrupt_excel",
        "notes": f"Cannot read as Excel; magic={file_type}",
    }


def find_guid_for_path(path: Path, manifest: list[dict]) -> str:
    path_str = str(path.relative_to(ROOT))
    for item in manifest:
        for dl in item.get("downloads", []):
            if dl.get("path") == path_str:
                return item.get("guid", "")
    return ""


def main() -> None:
    parser = argparse.ArgumentParser(description="Verifica archivos Excel descargados.")
    parser.add_argument("--force", action="store_true", help="Re-verificar aunque ya exista resultado.")
    args = parser.parse_args()

    config = load_config()
    logger = setup_logging("06_verify_excel_files")

    manifest: list[dict] = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    # Build path → guid index from manifest
    path_to_guid: dict[str, str] = {}
    for item in manifest:
        for dl in item.get("downloads", []):
            p = dl.get("path")
            if p:
                path_to_guid[p] = item.get("guid", "")

    # Load existing results for idempotency
    audit_json = AUDIT_DIR / "excel_file_audit.json"
    existing: dict[str, dict] = {}
    if audit_json.exists() and not args.force:
        try:
            for row in json.loads(audit_json.read_text(encoding="utf-8")):
                existing[row["path"]] = row
        except Exception:
            pass

    # Find all Excel files
    excel_files = sorted(
        list(ROOT.rglob("data/raw/**/*.xlsx")) + list(ROOT.rglob("data/raw/**/*.xls"))
    )
    logger.info("Archivos Excel encontrados: %s", len(excel_files))

    results: list[dict] = []
    for path in excel_files:
        rel = str(path.relative_to(ROOT))
        if rel in existing and not args.force:
            results.append(existing[rel])
            continue

        guid = path_to_guid.get(rel, find_guid_for_path(path, manifest))
        fmt_claimed = path.suffix.lower().lstrip(".")

        info = inspect_excel(path, logger)
        row = {
            "path": rel,
            "guid": guid,
            "format_claimed": fmt_claimed,
            **info,
        }
        results.append(row)
        logger.info("%s | %s | rows=%s | %s", path.name, info["actual_file_type"], info.get("rows_total"), info["status"])

    # Save JSON
    write_json(audit_json, results)

    # Save CSV
    csv_path = AUDIT_DIR / "excel_file_audit.csv"
    csv_cols = [
        "path", "guid", "format_claimed", "actual_file_type",
        "sha256", "sheet_count", "rows_total", "max_rows_single_sheet",
        "status", "notes",
    ]
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=csv_cols, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(results)

    # Stats
    by_status: dict[str, int] = {}
    for r in results:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1

    total_rows = sum(r.get("rows_total") or 0 for r in results)
    valid = sum(v for s, v in by_status.items() if "valid" in s)
    corrupt = sum(v for s, v in by_status.items() if s in ("corrupt_excel", "redirect_stub_unresolved", "html_error_file"))

    lines = [
        "# Auditoría de Archivos Excel",
        "",
        f"Fecha: 2026-04-28",
        f"Archivos evaluados: {len(results)}",
        f"Filas totales: {total_rows:,}",
        "",
        "## Resumen por estado",
        "",
        "| Estado | Cantidad |",
        "|--------|----------|",
    ]
    for status, count in sorted(by_status.items(), key=lambda x: -x[1]):
        lines.append(f"| `{status}` | {count} |")

    lines += [
        "",
        f"**Válidos**: {valid}  ",
        f"**Problemáticos**: {corrupt}  ",
        "",
        "## Archivos problemáticos",
        "",
    ]
    for r in results:
        if r["status"] not in ("xlsx_valid", "xls_valid"):
            lines.append(f"- `{r['path']}` ({r['guid']}) → {r['status']}: {r.get('notes', '')}")

    if valid == len(results):
        lines.append("\n✓ Todos los archivos Excel son válidos.")

    (AUDIT_DIR / "EXCEL_FILE_AUDIT.md").write_text("\n".join(lines) + "\n", encoding="utf-8")

    logger.info(
        "Verificación completa: %s archivos, %s válidos, %s problemáticos, %s filas totales",
        len(results), valid, corrupt, total_rows,
    )
    print(f"\nVerificados: {len(results)}, Válidos: {valid}, Problemáticos: {corrupt}")
    print(f"Filas totales en Excel: {total_rows:,}")
    print(f"Archivos: {csv_path} | {audit_json}")


if __name__ == "__main__":
    main()
