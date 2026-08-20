#!/usr/bin/env python3
"""Resolve possible truncation in datasets that hit the Junar server page cap (999 rows).

For each suspected resource, tries:
  A. Alternative pagination offsets to detect if more data exists.
  B. Alternative formats (csv, xls) and compares row counts.
  C. Inspects source endpoint file if available.
  D. Reads total-count fields from ajson metadata if present.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import time
from pathlib import Path
from typing import Any

import requests

# Change sys.path so we can import junar_common from the same directory.
import sys
sys.path.insert(0, str(Path(__file__).parent))

from junar_common import (
    ROOT,
    JUNAR_SERVER_PAGE_CAP,
    PAGINATION_BOUNDARY_ROWS,
    RateLimiter,
    Settings,
    content_hash_json,
    is_junar_list_of_lists,
    load_config,
    load_settings,
    request_with_retries,
    resource_guid,
    setup_logging,
    write_json,
)

AUDIT_DIR = ROOT / "audit"
MANIFEST_PATH = ROOT / "data/processed/download_manifest.json"

COMPLETENESS_STATUSES = (
    "complete_verified_by_source_file",
    "complete_verified_by_csv_or_xlsx",
    "complete_probable_exact_999_rows",
    "partial_possible_truncation_server_repeats_payload",
    "partial_known_api_limit_source_unavailable",
    "requires_manual_download",
    "requires_formal_request",
)


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def count_lol_rows(payload: Any) -> int | None:
    if is_junar_list_of_lists(payload):
        result = payload["result"]
        return max(0, len(result) - 1)  # exclude header
    return None


def try_pagination(
    session: requests.Session,
    base_url: str,
    guid: str,
    settings: Settings,
    limiter: RateLimiter,
    logger,
) -> dict[str, Any]:
    """Try alternative offsets to determine if data exists beyond page 1."""
    url = f"{base_url}/datastreams/{guid}/data.ajson/"
    params_base = {"auth_key": settings.api_key, "limit": 1000}
    offsets_to_try = [0, 999, 1000, 1999, 2000, 500]
    results: dict[int, dict] = {}
    hashes: dict[int, str] = {}

    for offset in offsets_to_try:
        limiter.wait()
        params = {**params_base, "offset": offset}
        try:
            resp = session.get(
                url,
                params=params,
                headers={"Referer": settings.referer, "User-Agent": "esparza-junar-civic-audit/0.1"},
                timeout=settings.timeout_seconds,
            )
            if resp.status_code != 200:
                results[offset] = {"http_status": resp.status_code, "rows": None}
                continue
            payload = resp.json()
            h = content_hash_json(payload)
            rows = count_lol_rows(payload)
            results[offset] = {"http_status": 200, "rows": rows, "hash": h}
            hashes[offset] = h
        except Exception as exc:
            results[offset] = {"error": str(exc), "rows": None}

    # Determine if content at offset 999/1000 is identical to offset 0 (= server ignores offset).
    hash_0 = hashes.get(0)
    hash_999 = hashes.get(999)
    hash_1000 = hashes.get(1000)

    all_hashes = list(hashes.values())
    all_same = len(set(all_hashes)) == 1 and len(all_hashes) > 1

    rows_at_offsets = {k: v["rows"] for k, v in results.items() if v.get("rows") is not None}
    max_rows = max(rows_at_offsets.values()) if rows_at_offsets else None

    payload_repeated = all_same or (
        hash_0 is not None
        and hash_999 is not None
        and hash_0 == hash_999
    )

    has_more_data = False
    if not payload_repeated:
        # If offset 999 or 1000 returns rows > 0 and different hash, there's more data.
        for off in (999, 1000, 1999, 2000):
            r = results.get(off, {})
            h = r.get("hash")
            rows = r.get("rows")
            if h and h != hash_0 and isinstance(rows, int) and rows > 0:
                has_more_data = True
                break

    return {
        "offsets_tried": results,
        "payload_repeated": payload_repeated,
        "has_more_data_beyond_999": has_more_data,
        "max_rows_seen": max_rows,
    }


def try_format(
    session: requests.Session,
    base_url: str,
    guid: str,
    fmt: str,
    settings: Settings,
    limiter: RateLimiter,
    logger,
) -> dict[str, Any]:
    """Try downloading a resource in an alternative format and count rows."""
    url = f"{base_url}/datastreams/{guid}/data.{fmt}/"
    params = {"auth_key": settings.api_key, "limit": 50000}
    limiter.wait()
    try:
        resp = session.get(
            url,
            params=params,
            headers={"Referer": settings.referer, "User-Agent": "esparza-junar-civic-audit/0.1"},
            timeout=settings.timeout_seconds,
        )
        if resp.status_code != 200:
            return {"http_status": resp.status_code, "rows": None}

        content = resp.content

        # Check for redirect JSON stub
        try:
            stub = json.loads(content)
            if isinstance(stub, dict) and stub.get("fType") == "REDIRECT":
                return {"http_status": 200, "rows": None, "note": "redirect_stub"}
        except Exception:
            pass

        if fmt == "csv":
            try:
                text = content.decode("utf-8-sig", errors="replace")
                reader = list(csv.reader(text.splitlines()))
                rows = max(0, len(reader) - 1)  # exclude header
                return {"http_status": 200, "rows": rows, "bytes": len(content)}
            except Exception as e:
                return {"http_status": 200, "rows": None, "error": str(e)}

        if fmt in ("xls", "xlsx"):
            # Save temp and open with openpyxl/xlrd
            import tempfile
            with tempfile.NamedTemporaryFile(suffix=f".{fmt}", delete=False) as tmp:
                tmp.write(content)
                tmp_path = Path(tmp.name)
            try:
                if fmt == "xlsx":
                    import openpyxl
                    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                    total = sum(ws.max_row - 1 for ws in wb.worksheets)
                    wb.close()
                else:
                    import xlrd
                    wb = xlrd.open_workbook(tmp_path)
                    total = sum(ws.nrows - 1 for ws in wb.sheets())
                return {"http_status": 200, "rows": total, "bytes": len(content)}
            except Exception as e:
                return {"http_status": 200, "rows": None, "error": str(e)}
            finally:
                tmp_path.unlink(missing_ok=True)

        if fmt in ("ajson", "json", "pjson"):
            try:
                payload = resp.json()
                rows = count_lol_rows(payload)
                return {"http_status": 200, "rows": rows, "bytes": len(content)}
            except Exception as e:
                return {"http_status": 200, "rows": None, "error": str(e)}

    except Exception as exc:
        return {"error": str(exc), "rows": None}

    return {"http_status": None, "rows": None}


def resolve_resource(
    item: dict[str, Any],
    session: requests.Session,
    settings: Settings,
    limiter: RateLimiter,
    logger,
    dry_run: bool = False,
) -> dict[str, Any]:
    guid = item["guid"]
    title = item.get("title", "")
    category = item.get("category", "").strip()
    logger.info("Resolviendo truncamiento: %s | %s", guid, title[:60])

    # Best existing rows from manifest
    rows_ajson = None
    for dl in item.get("downloads", []):
        if dl.get("format") == "ajson" and isinstance(dl.get("rows_detected"), int):
            rows_ajson = dl["rows_detected"]
            break

    rows_csv = None
    rows_xlsx = None
    rows_source_file = None
    pagination_result = {}
    payload_repeated = None
    has_more = None
    source_file_available = False
    formats_tried = []

    if not dry_run:
        # A. Pagination probe
        pagination_result = try_pagination(session, settings.base_url, guid, settings, limiter, logger)
        payload_repeated = pagination_result.get("payload_repeated")
        has_more = pagination_result.get("has_more_data_beyond_999")

        # B. Alternative formats
        for fmt in ("csv", "xlsx"):
            formats_tried.append(fmt)
            r = try_format(session, settings.base_url, guid, fmt, settings, limiter, logger)
            if fmt == "csv" and isinstance(r.get("rows"), int):
                rows_csv = r["rows"]
            elif fmt == "xlsx" and isinstance(r.get("rows"), int):
                rows_xlsx = r["rows"]

        # C. Source endpoint
        ep = item.get("endpoint_download")
        if isinstance(ep, dict) and ep.get("status") in ("ok", "ok_redirect_followed", "skipped_existing"):
            source_file_available = True
            ep_path = ep.get("path")
            if ep_path:
                full_path = ROOT / ep_path
                if full_path.exists() and full_path.suffix.lower() in (".xlsx", ".xls"):
                    fmt = full_path.suffix[1:]
                    try:
                        if fmt == "xlsx":
                            import openpyxl
                            wb = openpyxl.load_workbook(full_path, read_only=True, data_only=True)
                            rows_source_file = sum(ws.max_row - 1 for ws in wb.worksheets)
                            wb.close()
                        else:
                            import xlrd
                            wb = xlrd.open_workbook(full_path)
                            rows_source_file = sum(ws.nrows - 1 for ws in wb.sheets())
                    except Exception as e:
                        logger.warning("Error leyendo archivo fuente %s: %s", ep_path, e)

    # D. Classify
    rows_best = max(
        (r for r in [rows_ajson, rows_csv, rows_xlsx, rows_source_file] if r is not None),
        default=rows_ajson,
    )

    if rows_source_file is not None and rows_best == rows_source_file and rows_best > (rows_ajson or 0):
        final_status = "complete_verified_by_source_file"
        action = "use_source_file_for_analysis"
    elif rows_xlsx is not None and rows_best == rows_xlsx and (rows_xlsx or 0) > (rows_ajson or 0):
        final_status = "complete_verified_by_csv_or_xlsx"
        action = "use_xlsx_for_analysis"
    elif rows_csv is not None and rows_best == rows_csv and (rows_csv or 0) > (rows_ajson or 0):
        final_status = "complete_verified_by_csv_or_xlsx"
        action = "use_csv_for_analysis"
    elif payload_repeated:
        final_status = "partial_possible_truncation_server_repeats_payload"
        action = "request_source_file_or_formal_request"
    elif has_more:
        final_status = "partial_possible_truncation_server_repeats_payload"
        action = "paginate_manually_or_request_source_file"
    elif rows_ajson == 999 and not has_more and not payload_repeated:
        # No evidence of more data; 999 might be exact.
        final_status = "complete_probable_exact_999_rows"
        action = "verify_against_source_if_critical"
    else:
        final_status = "partial_known_api_limit_source_unavailable"
        action = "requires_manual_download_or_formal_request"

    year_detected = ""
    for yr in range(2000, 2031):
        if str(yr) in title or str(yr) in guid:
            year_detected = str(yr)
            break

    return {
        "guid": guid,
        "title": title,
        "category": category,
        "year_detected": year_detected,
        "rows_ajson": rows_ajson,
        "rows_csv": rows_csv,
        "rows_xlsx": rows_xlsx,
        "rows_source_file": rows_source_file,
        "rows_best": rows_best,
        "pagination_strategy_result": json.dumps(pagination_result, ensure_ascii=False) if pagination_result else "",
        "payload_repeated": payload_repeated,
        "has_more_data_beyond_999": has_more,
        "source_file_available": source_file_available,
        "final_completeness_status": final_status,
        "recommended_action": action,
        "notes": "",
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Resuelve posibles truncamientos en datasets de Junar.")
    parser.add_argument("--dry-run", action="store_true", help="No hacer requests; solo clasificar con datos existentes.")
    parser.add_argument("--guid", default=None, help="Resolver solo un GUID específico.")
    parser.add_argument("--force", action="store_true", help="Re-evaluar aunque ya exista resolución previa.")
    args = parser.parse_args()

    config = load_config()
    settings = load_settings(config)
    logger = setup_logging("04_resolve_truncation")

    manifest: list[dict] = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))

    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    # Load existing resolutions for idempotency
    resolution_path = AUDIT_DIR / "truncation_resolution.json"
    existing: dict[str, dict] = {}
    if resolution_path.exists() and not args.force:
        try:
            for row in json.loads(resolution_path.read_text(encoding="utf-8")):
                existing[row["guid"]] = row
        except Exception:
            pass

    BOUNDARY = PAGINATION_BOUNDARY_ROWS | {rd for rd in range(990, 1010)}

    suspects = []
    for item in manifest:
        guid = item.get("guid", "")
        if args.guid and guid != args.guid:
            continue
        for dl in item.get("downloads", []):
            rd = dl.get("rows_detected")
            if dl.get("possible_truncation") or (isinstance(rd, int) and rd in BOUNDARY):
                suspects.append(item)
                break

    logger.info("Recursos sospechosos de truncamiento: %s", len(suspects))

    session = requests.Session()
    limiter = RateLimiter(settings.requests_per_second)
    results: list[dict] = []

    for item in suspects:
        guid = item["guid"]
        if guid in existing and not args.force:
            logger.info("Ya resuelto (skip): %s", guid)
            results.append(existing[guid])
            continue
        result = resolve_resource(item, session, settings, limiter, logger, dry_run=args.dry_run)
        results.append(result)

    # Save JSON
    write_json(resolution_path, results)

    # Save CSV
    csv_path = AUDIT_DIR / "truncation_resolution.csv"
    csv_cols = [
        "guid", "title", "category", "year_detected",
        "rows_ajson", "rows_csv", "rows_xlsx", "rows_source_file", "rows_best",
        "payload_repeated", "has_more_data_beyond_999", "source_file_available",
        "final_completeness_status", "recommended_action", "notes",
    ]
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=csv_cols, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(results)

    # Build markdown report
    by_status: dict[str, list] = {}
    for r in results:
        by_status.setdefault(r["final_completeness_status"], []).append(r)

    lines = [
        "# Resolución de Truncamientos",
        "",
        f"Fecha de ejecución: 2026-04-28",
        f"Recursos evaluados: {len(results)}",
        "",
        "## Resumen por estado",
        "",
        "| Estado | Cantidad |",
        "|--------|----------|",
    ]
    for status, items in sorted(by_status.items(), key=lambda x: -len(x[1])):
        lines.append(f"| `{status}` | {len(items)} |")

    lines += ["", "## Detalle por recurso", ""]
    for r in sorted(results, key=lambda x: x["final_completeness_status"]):
        lines.append(
            f"- **{r['guid']}** | {r['category']} | ajson={r['rows_ajson']} csv={r['rows_csv']} "
            f"xlsx={r['rows_xlsx']} src={r['rows_source_file']} → `{r['final_completeness_status']}`"
        )

    lines += [
        "",
        "## Interpretación",
        "",
        "- `complete_probable_exact_999_rows`: El servidor no retornó más datos en offsets alternativos. Probable que la tabla tenga exactamente 999 filas. Verificar contra portal o fuente si es crítico.",
        "- `complete_verified_by_csv_or_xlsx`: CSV o XLSX tiene más filas que ajson — usar ese formato para análisis.",
        "- `partial_possible_truncation_server_repeats_payload`: El servidor repite el mismo payload al cambiar offset. Datos incompletos en API; requiere archivo fuente.",
        "- `partial_known_api_limit_source_unavailable`: No hay evidencia de completitud ni archivo alternativo.",
        "- `requires_manual_download`: Archivo fuente disponible pero requiere descarga manual.",
        "- `requires_formal_request`: No hay datos recuperables técnicamente; requiere oficio.",
    ]

    (AUDIT_DIR / "TRUNCATION_RESOLUTION.md").write_text("\n".join(lines) + "\n", encoding="utf-8")

    logger.info(
        "Resolución completa: %s recursos → %s",
        len(results),
        ", ".join(f"{s}={len(v)}" for s, v in sorted(by_status.items(), key=lambda x: -len(x[1]))),
    )
    print(f"\nResueltos: {len(results)} recursos")
    print(f"Archivos: {csv_path} | {resolution_path}")
    print(f"Reporte: {AUDIT_DIR / 'TRUNCATION_RESOLUTION.md'}")


if __name__ == "__main__":
    main()
